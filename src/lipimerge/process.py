from pathlib import Path
from datetime import datetime, timezone
import openpyxl.workbook.workbook as workbook
from lipimerge import __version__
import lipimerge.internal.console as console
from lipimerge.internal.utils import selection_sort_step
import lipimerge.internal.xutils as xutils
import lipimerge.internal.exceptions as lipiex

def process(input_files: list[str], output_file: str, ignore_blanks: bool, trim_class_names: bool) -> lipiex.Success:
    """
    Process the input Excel files and save the merged result to the output file.
    
    Args:
        input_files (list[str]): List of input Excel file paths.
        output_file (str): Path to the output Excel file.
    
    Raises:
        LipiMergeException | *
    """

    # INIT

    result = workbook.Workbook()
    result.active.title = "lipimerge.log"
    log_sheet = result.active

    # ! adjust `logx*` if you chage this log_sheet initialization
    log_sheet.append([f"LipidQuant merge v. {__version__}"])
    log_sheet.append([f"Time (Local)", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    log_sheet.append([f"Time (UTC)", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S %Z%z")])
    log_sheet.append([])
    log_sheet.append(["File", "Sheet", "Status"])

    def logx(message: str):
        log_sheet.cell(row=log_sheet.max_row, column=3, value=message)

    def logxok():
        for row in log_sheet.iter_rows(min_row=6):
            row[2].value = "OK"

    # CHECK

    if Path(output_file).exists():
        raise lipiex.InvalidFile("Output file already exists.", output_file, ["Please delete the file and try again."])
    
    for i in range(len(input_files)):
        if input_files[i] in input_files[i+1:]:
            raise lipiex.InvalidFile("Duplicate input files.", input_files[i], ["The same file must not be merged twice."])
    
    inconsistencies = []
    ctx = {}
    for input_file in console.track(input_files, "Validating input files: "):
        with xutils.workbook(input_file) as wb:
            inconsistencies = xutils.validate_input(wb, input_file, ignore_blanks, inconsistencies, ctx)

    if inconsistencies:
        for c in inconsistencies:
            log_sheet.append([c.workbook, c.sheet, f"[R{c.row}C{c.col}]: {c.message}"])
        result.save(output_file)

        raise lipiex.ConsistencyError("Invalid input files.", ["Log has been saved into the output file."])
    
    # DOIT

    for input_file in console.track(input_files, "Merging: "):
        with xutils.workbook(input_file) as wb:
            for sheet in wb.worksheets:

                sheet_name = sheet.title
                log_sheet.append([input_file, sheet_name, "?"])
                if not sheet_name in result.sheetnames: result.create_sheet(sheet_name)

                logx("Processing: ")
                xutils.merge(result[sheet_name], sheet, ignore_blanks)
                logx("Waiting for validation!")
    
    # SORT & SAVE

    if trim_class_names:
        for sheet_name in result.sheetnames:
            if sheet_name == 'lipimerge.log': continue

            sheet = result[sheet_name]
            for col in sheet.iter_cols(min_col=2):
                if col[0].value is None: continue
                col[0].value = col[0].value.strip()
                
    for sh, sheet_name in enumerate(result.sheetnames):
        if sheet_name == 'lipimerge.log': continue
        
        sheet = result[sheet_name]
        for i in console.track(range(2, sheet.max_column), f"Sorting the merge ({sh}/{len(result.sheetnames)-1}): "):
        # `-1` counts for the 'lipimerge.log' sheet,
        # which is always present and is always the first in the `result` workbook
            selection_sort_step(sheet, less=xutils.column_less, swap=xutils.column_swap, end=sheet.max_column, begin=i) 
    
    console.print(f"Saving [{output_file}] ...")
    result.save(output_file)
    console.print()

    # VALIDATE

    with xutils.workbook(output_file) as out:

        for input_file in console.track(input_files, "Output validation (1/2): "):
            with xutils.workbook(input_file) as inp:
                for sheet in inp.worksheets:
                    if not sheet.title in out.sheetnames:
                        raise lipiex.ConsistencyError(
                            "Validation failed.",
                            [f"Missing sheet: '{sheet.title}'"]
                        )
                    xutils.clear_found_values(out[sheet.title], sheet, ignore_blanks, trim_class_names)

        for sheet_name in console.track(out.sheetnames, "Output validation (2/2): "):
            if sheet_name == 'lipimerge.log': continue
            if not xutils.has_empty_data_set(out[sheet_name]):
                raise lipiex.ConsistencyError(
                    "Validation failed.",
                    [f"Sheet: '{sheet_name}'"]
                )
            
    # FINALIZE

    logxok()
    console.print(f"Finalizing [{output_file}] ...")
    result.save(output_file)
                
    return lipiex.Success(
        f"{len(input_files)} files successfully merged into '{output_file}'.",
        input_files
    )

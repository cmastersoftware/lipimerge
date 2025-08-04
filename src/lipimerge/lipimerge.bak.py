# IMPORTS

# .. Global

import sys
from openpyxl import load_workbook

# .. Local

# MAIN


# Bubble sort >

# Bubble sort columns in an Excel sheet based on the first row values
# This script sorts columns in an Excel sheet by comparing the values in the first row of each column.
# It uses a bubble sort algorithm to swap columns until they are sorted in ascending order. 
from openpyxl import load_workbook

def get_column_header(ws, col):
    """Returns the value in the first row of the given column."""
    return ws.cell(row=1, column=col).value

def swap_columns(ws, col1, col2):
    """Swaps all cell values between two columns."""
    for row in range(1, ws.max_row + 1):
        cell1 = ws.cell(row=row, column=col1)
        cell2 = ws.cell(row=row, column=col2)
        cell1.value, cell2.value = cell2.value, cell1.value

# Load workbook and active sheet
wb = load_workbook("your_file.xlsm")
ws = wb.active

# Bubble sort columns based on first-row values
for i in range(1, ws.max_column):
    for j in range(1, ws.max_column - i + 1):
        val_j = get_column_header(ws, j)
        val_j1 = get_column_header(ws, j + 1)

        # Sort by string comparison (adjust as needed for numeric or case-insensitive)
        if str(val_j) > str(val_j1):
            swap_columns(ws, j, j + 1)

# Save the result
wb.save("sorted_columns_inplace.xlsm")

# < Bubble sort

def main():
    wb = load_workbook('Data/TestFile-1.xlsx', data_only=True)
    sheet = wb['Conc']  # or wb.active

    sheet.append([1, 2, 3])  # Example of appending a row

    q = next(sheet.rows)
    q = [cell.value for cell in q]
    print(q)

    # Iterate over rows and print plain cell values
    for row in sheet.iter_rows(values_only=True):
        for cell in row:
            if cell is not None:
                print(cell, end=' ')
            else:
                print('None', end=' ')
        print()

if __name__ == "__main__":
    main()

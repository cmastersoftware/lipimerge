from math import isclose
from contextlib import contextmanager
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.workbook.workbook import Workbook
from lipimerge.internal.utils import quick_sort
from lipimerge.internal.utils import is_blank
import lipimerge.internal.exceptions as lipiex

def _dump(sheet: Worksheet) -> str:
    return str([row for row in sheet.iter_rows(values_only=True)])


@contextmanager
def workbook(filepath: str):
    """
    A context manager for openpyxl workbooks, which ensures the workbook is closed after use.

    Args:
        filepath (str): The path to the Excel file.

    Yields:
        openpyxl.workbook.workbook.Workbook: The loaded or new workbook object.

    Raise:
        re-raise exception when opening or closing the workbook

    Note:
       Opens the workbook with default settings. see openpyxl.load_workbook
    """

    wb = None
    try:
        wb = load_workbook(filepath)
        yield wb  # Yield the workbook object to the 'with' block

    except Exception:
        raise

    finally:
        if wb is not None: wb.close()


def get_class_names(sheet: Worksheet) -> list[str|None]:
    """
    Return lipid class names in the LipidQuant Excel sheet
    
    The lipid class names are values in the first row of the worksheet.

    Args:
        sheet (Worksheet): A LipidQuant Excel sheet
    
    Returns:
        list[str|None]: values in the first row of the `sheet`
                        The list contains `None` at positions of empty cells
                        If the `sheet` is empty, the list contains a single `None` value.
    """
    # The construct is surprisingly nontrivial, thats why encapsulated in this helper function.
    # Note: `next(sheet.rows)` raises with empty sheets (even though iter_rows does not)
    return [value for value in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]


def get_record_names(sheet: Worksheet) -> list[str|None]:
    """
    Returns record names in the LipidQuant Excel sheet

    The record names are values in the first column of the worksheet.

    Args:
        sheet (Worksheet): A LipidQuant Excel sheet
    
    Returns:
        list[str|None]: values in the first column of the `sheet`
                        The list contains `None` at positions of empty cells
                        If the `sheet` is empty, the list contains a single `None` value.
                        The empty cell A1 **is** included
    """
    # see get_class_names
    return [value for value in next(sheet.iter_cols(min_col=1, max_col=1, values_only=True))]


def equal(sheet1: Worksheet, sheet2: Worksheet) -> bool:
    """
    Check if two worksheets are equal.

    Used for testing purposes.

    Args:
        sheet1 (Worksheet): An openpyxl Worksheet object.
        sheet2 (Worksheet): An openpyxl Worksheet object.

    Returns:
        bool: True if the sheets are equal.
    """
    if sheet1.max_row != sheet2.max_row or sheet1.max_column != sheet2.max_column:
        return False

    for irow in range(1, sheet1.max_row + 1):
        for icol in range(1, sheet1.max_column + 1):
            if sheet1.cell(irow, icol).value != sheet2.cell(irow, icol).value:
                return False

    return True


class InputInconsistency:
    def __init__(self, message: str, workbook: str, sheet: str, row: int, col: int):
        self.message = message
        self.workbook = workbook
        self.sheet = sheet
        self.row = row
        self.col = col
    
    def __str__(self):
        return f"'[{self.workbook}]{self.sheet}'!R{self.row}C{self.col}: {self.message}"

def validate_input(workbook: Workbook, name: str, ignore_blanks: bool, result: list[InputInconsistency], context: dict[tuple[str, str], str]) -> list[InputInconsistency]:
    """
    Check for possible inconsistencies in a LipidQuant Excel workbook.
    
    Validity constraints:
    - For every sheet in the workbool:
        - the sheet is empty or the first cell (A1) is empty
        - no duplicate lipid class names (1st row cells)
        - for each column starting from the second column, and
          for each row starting from the second row:
            - If a class name (the cell in the 1st row) is empty, all cells in that column must be empty.
            - If a record name (the cell in the 1st column) is empty, all cells in that row must be empty.
        - blank must not contain data
    - For any two workbooks (including the workbook itself):
        - Two sheets of the same name do not have the same record name (1st column cells)
          and class name (1st row cells) combination.

    Args:
        workbook (Workbook): A LipidQuant Excel workbook.
        name (str): The name of the workbook (file name; for error messages).
        result (list[InputInconsistency]): Output parameter: A list of found inconsistencies.
                                           The function shall be called with an empty list for the first call,
                                           and the list shall be passed to subsequent calls.
                                           It gets populated with the found inconsistencies or remains empty
                                           if none is found.
        ignore_blanks (bool): if True, blank samples are excluded from the validation.
        context (dict[tuple[str, str, str], str]): An opaque context for cross-workbook validation.
                                                   The function shall be called with an empty dict for the first call,
                                                   and the dict shall be passed to subsequent calls.
    
    Returns:
        list[InputInconsistency]: The `result`
    """

    # I originally considered making sheets with less than 2 rows or cols invalid,
    # but such sheet may result from merge if a source sheet has all (but the first) rows empty
    # or all (but the first) columns empty.

    # INIT

    def add(message: str, workbook: str, sheet: str, row: int, col: int):
        result.append(InputInconsistency(message=message, workbook=workbook, sheet=sheet, row=row, col=col))

    def validate_sheet(sheet: Worksheet):

        # .. is lipimerge sheet
        if sheet.cell(1,1).value is not None: 
            add("Non-empty cell A1 (skipping the rest of the sheet)", name, sheet.title, 1, 1)
            return result
        
        # .. has no duplicate classes
        cls_names = get_class_names(sheet)
        already_found = []
        for i, cls in enumerate(cls_names):
            if cls is None: continue
            if cls in already_found: continue

            for j in range(i+1, len(cls_names)):
                if cls_names[j] == cls:
                    if not cls in already_found:
                        add(f"Duplicate class '{cls}' within one data sheet.", name, sheet.title, 1, i+1)
                        already_found.append(cls)
                    add(f"Duplicate class '{cls}' within one data sheet.", name, sheet.title, 1, j+1)
                
        # .. has no duplicate records
        rec_names = get_record_names(sheet)
        already_found = []
        for i, rec in enumerate(rec_names):
            if rec is None: continue
            if rec in already_found: continue
            if ignore_blanks and is_blank(rec): continue

            for j in range(i+1, len(rec_names)):
                if rec_names[j] == rec:
                    if not rec in already_found:
                        add(f"Duplicate record '{rec}' within one data sheet.", name, sheet.title, i+1, 1)
                        already_found.append(rec)
                    add(f"Duplicate record '{rec}' within one data sheet.", name, sheet.title, j+1, 1)

        # .. all blanks have no values
        if not ignore_blanks:
            for i, rec in enumerate(rec_names):
                if not is_blank(rec): continue

                irow = i+1
                for icol in range(2, sheet.max_column):
                    if sheet.cell(irow, icol).value is not None:
                        add("Non-empty cell in a blank.", name, sheet.title, irow, icol)


    def validate_empty_cells(sheet: Worksheet):

        for icol in range(2, sheet.max_column + 1):
            if sheet.cell(1, icol).value is not None: continue
            for irow in range(2, sheet.max_row + 1):
                if sheet.cell(irow, icol).value is not None:
                    add("Non-empty cell in a column with an empty class name.", name, sheet.title, irow, icol)

        for irow in range(2, sheet.max_row + 1):
            if sheet.cell(irow, 1).value is not None: continue
            for icol in range(2, sheet.max_column + 1):
                if sheet.cell(irow, icol).value is not None:
                    add("Non-empty cell in a row with an empty record.", name, sheet.title, irow, icol)

    def validate_duplicates_across_workbooks(sheet: Worksheet):

        for i, record in enumerate(get_record_names(sheet)):
            if record is None: continue
            if ignore_blanks and is_blank(record): continue

            for j, cls in enumerate(get_class_names(sheet)):
                if cls is None: continue

                key = (sheet.title, record, cls)
                rec = context.get(key, None)
                if rec is None:
                    context[key] = (name, i+1, j+1)
                else:
                    if rec[0] == name: continue # the same workbook; already reported in `validate_sheet`
                    add(f"Duplicate class '{cls}' and record '{record}' within one data sheet across workbooks.", rec[0], sheet.title, rec[1], rec[2])
                    add(f"Duplicate class '{cls}' and record '{record}' within one data sheet across workbooks.", name, sheet.title, i+1, j+1)

    # DOIT

    for sheet in workbook.worksheets:

        validate_sheet(sheet)
        validate_empty_cells(sheet)
        validate_duplicates_across_workbooks(sheet)

    # DONE

    return result


def merge(destination: Worksheet, source: Worksheet, ignore_blanks: bool) -> Worksheet:
    """
    Merges source into destination.

    The merge operation appends rows from `source` to `destination`.
    See README.md for details.

    Args:
        destination (Worksheet): A valid (see `validate_input`) LipidQuant Excel sheet into which the `source` is to be merged
        source (Worksheet): A valid (see `validate_input`) LipidQuant Excel sheet to merge into `destination`.
        ignore_blanks (bool): If True, blank sample records are excluded from the merge

    Returns:
        Worksheet: `destination`

    Notes:
        - `destination` will be modified with the merge (must be writable)
        - the merge operation preserve valid LipidQuant format, so consecutive merges into `destination` are supported
        - the merge operation guarantees that no empty rows or columns remain in the merge (they are ignored in the `source`)
        - the merge operation can be initiated with an empty `destination` sheet
        - `source` remains unmodified (can be read only)
    """
    
    dst_classes = get_class_names(destination)
    src_classes = get_class_names(source)
    dst_records = get_record_names(destination)

    def mapper(destination: Worksheet, dst_classes: list[str|None], value: str) -> int|None:
        if str is None: return None
        try:
            # openpyxl Excel cells indexing starts from 1
            # while Python indexing is zero-based.
            # We want this function to obey Excel indexing
            return dst_classes.index(value) + 1 
        except:
            destination.cell(1, destination.max_column+1, value)
            return destination.max_column

    # mapper may already modify destination, see above
    src_dst_map = [mapper(destination, dst_classes, value) for value in src_classes]

    for row in source.iter_rows(2, values_only=True):
        if row[0] is None: continue
        if ignore_blanks and is_blank(row[0]): continue

        try:
            irow = dst_records.index(row[0]) + 1
        except ValueError:
            destination.append([row[0]])
            irow = destination.max_row

        for icol in range(1,len(src_dst_map)):
            if src_dst_map[icol] is None: continue
            destination.cell(irow, src_dst_map[icol], row[icol])

    return destination


def column_less(sheet: Worksheet, lhs: int, rhs: int) -> bool:
    """
    Compare two columns by their lipid class names.
    (see get_class_names)

    Args:
        sheet (Worksheet): An openpyxl Worksheet object.
        lhs (int): left-hand-side column index (1-based)
        rhs (int): right-hand-side column index (1-based)
    
    Returns:
        bool: True if lhs < rhs

    Warning:
        The `sheet` must contain no empty columns (see `merge`).
    """
    return sheet.cell(1, lhs).value < sheet.cell(1, rhs).value


def column_swap(sheet: Worksheet, i: int, j: int) -> Worksheet:
    """
    Swaps two columns in the sheet.

    In-place swap by swapping every cell value.

    Args:
        sheet (Worksheet): An openpyxl Worksheet object.
        i (int): index of one column to swap (1-based)
        j (int): index of the other column to swap (1-based)

    Returns:
        Worksheet: `sheet` after swap

    Raise:
        InvalidCellIndex
        # swapping in an empty sheet is a no-op; never raise if i, j > 0
    """
    if (

        i > 0 and j > 0 and
        sheet.max_column == 1 and sheet.max_row == 1 and
        sheet.cell(1, 1).value is None

    ): return sheet
        

    if i < 1 or j < 1:
        raise lipiex.InvalidCellIndex("column_swap", min(i, j), 1, sheet.max_column, sheet.max_row)
    if i > sheet.max_column or j > sheet.max_column:
        raise lipiex.InvalidCellIndex("column_swap", max(i, j), 1, sheet.max_column, sheet.max_row)

    if i == j: return sheet

    for row in sheet.iter_rows(values_only=False):
        row[i-1].value, row[j-1].value = row[j-1].value, row[i-1].value

    return sheet


def column_sort(sheet: Worksheet) -> Worksheet:
    """
    Sorts columns in the sheet acc. their lipid class names.
    (see column_less)
    (see column_swap)

    Sorting starts from column 'B' (the first column 'A' is reserved for record names).

    Args:
        sheet (Worksheet): A valid (see validate_input) LipidQuant Excel sheet
    
    Retrurns:
        Worksheet: `sheet` after sort
    """
    raise NotImplementedError("`column_sort` implementation deprecated in favor of displaying the progress bar")
    return quick_sort(sheet, less=column_less, swap=column_swap, end=sheet.max_column, begin=2)

    
def clear_found_values(destination: Worksheet, source: Worksheet, ignore_blanks: bool, trim_class_names: bool) -> Worksheet:
    """
    Check if a value for every record and lipid class in `source` exists in `destination`,
    and if so, clear that value (set it to `None`) in destination.

    The function does not simply search for a value, but it searches for a given
    data record name and lipid class name, and compares the values in `source` and `destination`.

    The idea is that if a sheet M resulted as a merge of sheets A, B and C,
    the consecutive calls to `clear_found_values(M, A)`, `...(M, B)`, `...(M, C)`
    shall result in an empty data set (containing only record names and lipid class names).

    Args:
        destination (Worksheet): A valid (see validate_input) LipidQuant Excel sheet
        source (Worksheet): A valid (see validate_input) LipidQuant Execl sheet to compare to `destination`
        ignore_blanks (bool): If True, blank sample records are excluded from the check and clearing
        trim_class_names (bool): If True, white spaces at the beginning and the end of a class name are trimmed

    Return:
        Worksheet: `destination`

    Raise:
        Exception: if data entry in `source` is not found in `destination` or the values differ

    Warning:
        `destination` gets modified (do not save it...) 
    """
    dst_records = get_record_names(destination)
    src_records = get_record_names(source)
    dst_classes = get_class_names(destination)
    src_classes = get_class_names(source)

    if ignore_blanks:
        for i in range(len(src_records)):
            if is_blank(src_records[i]): src_records[i] = None

    def mapper(dst_items: list[str|None], value: str) -> int|None:
        if str is None: return None
        
        try:
            # openpyxl Excel cells indexing starts from 1
            # while Python indexing is zero-based.
            # We want this function to obey Excel indexing
            return dst_items.index(value) + 1
        except:
            raise Exception(f"{value}")
        
    def trim_class_name(value: str|None) -> str|None:
        if value is None: return value
        return value.strip()

    try:
        src_dst_record_map = [mapper(dst_records, value) for value in src_records]
    except Exception as e:
        raise lipiex.ConsistencyError("Source data not found in destination.", details=[f"Data record: {str(e)}"])
    
    try:
        src_dst_classes_map = [
            mapper(
                dst_classes, 
                value if not trim_class_names else trim_class_name(value)
            ) 
            for value in src_classes
        ]
    except Exception as e:
        raise lipiex.ConsistencyError("Source data not found in destination.", details=[f"Lipid class: {str(e)}"])
    
    for src_row in range(2, source.max_row + 1):
        if source.cell(src_row, 1).value is None: continue
        if ignore_blanks and is_blank(source.cell(src_row, 1).value): continue

        for src_col in range(2, source.max_column + 1):
            src_val = source.cell(src_row, src_col).value
            if src_val is None: continue

            # We index openpyxl cells 1-base, but python list maps 0-based
            dst_row = src_dst_record_map[src_row - 1]
            dst_col = src_dst_classes_map[src_col - 1]
            dst_val = destination.cell(dst_row, dst_col).value

            if dst_val is None:
                raise lipiex.ConsistencyError(
                    "Source data not found in destination.",
                    details=[
                        f"Source record: {source.cell(src_row, 1).value} (row {src_row})",
                        f"Source lipid class: {source.cell(1, src_col).value} (column {src_col})",
                        f"Source value: {src_val}",
                        f"Destination record: {destination.cell(dst_row, 1).value} (row {dst_row})",
                        f"Destination lipid class: {destination.cell(1, dst_col).value} (column {dst_col})",
                        f"Destination value: [empty cell]"
                    ]
                )

            if not isclose(src_val, dst_val, rel_tol=1e-12, abs_tol=1e-12):
                raise lipiex.ConsistencyError(
                    "Source data differs for destination.",
                    details=[
                        f"Source record: {source.cell(src_row, 1).value} (row {src_row})",
                        f"Source lipid class: {source.cell(1, src_col).value} (column {src_col})",
                        f"Source value: {src_val}",
                        f"Destination record: {destination.cell(dst_row, 1).value} (row {dst_row})",
                        f"Destination lipid class: {destination.cell(1, dst_col).value} (column {dst_col})",
                        f"Destination value: {dst_val}",
                    ]
                )

            destination.cell(dst_row, dst_col).value = None

    return destination


def has_empty_data_set(sheet: Worksheet) -> bool:
    """
    Check whether all values in the dataset are empty cells.
    (see clear_found_values)

    Args:
        sheet (Worksheet): A valid (see validate_input) LilpidQuant Excel sheet
    
    Return:
        bool: True if the sheet is empty up to the record names and lipid class names
              (see get_record_names)
              (see get_class_names)
    """
    for row in sheet.iter_rows(min_row=2, values_only=True):
        for col in row[1:]: # row index is now 0-based
            if col is not None: return False

    return True

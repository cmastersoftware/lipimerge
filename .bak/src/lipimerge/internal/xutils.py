from contextlib import contextmanager
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from lipimerge.internal.utils import quick_sort
import lipimerge.internal.exceptions as lipiex

def _dump(sheet: Worksheet) -> str:
    return str([row for row in sheet.iter_rows(values_only=True)])


@contextmanager
def workbook(filepath: str):
    """
    A context manager for openpyxl workbooks, which ensures the workbook is closed after use.

    Args:
        filepath (str): The path to the Excel file.

    Yields:
        openpyxl.workbook.workbook.Workbook: The loaded or new workbook object.

    Raise:
        re-raise exception when opening or closing the workbook

    Note:
       Opens the workbook with default settings. see openpyxl.load_workbook
    """

    wb = None
    try:
        wb = load_workbook(filepath)
        yield wb  # Yield the workbook object to the 'with' block

    except Exception:
        raise

    finally:
        if wb is not None: wb.close()


def get_class_names(sheet: Worksheet) -> list[str|None]:
    """
    Return lipid class names in the LipidQuant Excel sheet
    
    The lipid class names are values in the first row of the worksheet.

    Args:
        sheet (Worksheet): A LipidQuant Excel sheet
    
    Returns:
        list[str|None]: values in the first row of the `sheet`
                        The list contains `None` at positions of empty cells
                        If the `sheet` is empty, the list contains a single `None` value.
    """
    # The construct is surprisingly nontrivial, thats why encapsulated in this helper function.
    # Note: `next(sheet.rows)` raises with empty sheets (even though iter_rows does not)
    return [value for value in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]


def get_record_names(sheet: Worksheet) -> list[str|None]:
    """
    Returns record names in the LipidQuant Excel sheet

    The record names are values in the first column of the worksheet.

    Args:
        sheet (Worksheet): A LipidQuant Excel sheet
    
    Returns:
        list[str|None]: values in the first column of the `sheet`
                        The list contains `None` at positions of empty cells
                        If the `sheet` is empty, the list contains a single `None` value.
    """
    # see get_class_names
    return [value for value in next(sheet.iter_cols(min_col=1, max_col=1, values_only=True))]


def equal(sheet1: Worksheet, sheet2: Worksheet) -> bool:
    """
    Check if two worksheets are equal.

    Used for testing purposes.

    Args:
        sheet1 (Worksheet): An openpyxl Worksheet object.
        sheet2 (Worksheet): An openpyxl Worksheet object.

    Returns:
        bool: True if the sheets are equal.
    """
    if sheet1.max_row != sheet2.max_row or sheet1.max_column != sheet2.max_column:
        return False

    for irow in range(1, sheet1.max_row + 1):
        for icol in range(1, sheet1.max_column + 1):
            if sheet1.cell(irow, icol).value != sheet2.cell(irow, icol).value:
                return False

    return True


class InputInconsistency:
    def __init__(self, message: str, row: int, col: int):
        self.message = message
        self.row = row
        self.col = col
    
    def __str__(self):
        return f"{self.message}: [{self.row}, {self.col}]"

def validate_input(sheet: Worksheet) -> list[InputInconsistency]:
    """
    Check for possible inconsistencies in a LipidQuant Excel sheet.
    
    Validity constraints:
    - the sheet is empty or the first cell (A1) is empty
    - no duplicate lipid class names (1st row cells)
    - no duplicate data record names (1st column row)
    - for each column starting from the second column, and
      for each row starting from the second row:
        - If a class name (the cell in the 1st row) is empty, all cells in that column must be empty.
        - If a record name (the cell in the 1st column) is empty, all cells in that row must be empty.

    Args:
        sheet (Worksheet): A LipidQuant Excel sheet
    
    Returns:
        list[InputInconsistency]: A list of found inconsistencies.
                                  Empty list if valid.
    """

    # I originally considered making sheets with less than 2 rows or cols invalid,
    # but such sheet may result from merge if a source sheet has all (but the first) rows empty
    # or all (but the first) columns empty.

    result = []

    def add(message: str, row: int, col: int):
        result.append(InputInconsistency(message=message, row=row, col=col))

    if sheet.cell(1,1).value is not None: add("Non-enpty cell A1", 1, 1)

    # classes and records always consist of at least a single `None` item
    
    classes = get_class_names(sheet)
    for i in range(0, len(classes)-1):
        if classes[i] is None: continue
        for j in range(i+1, len(classes)):
            if classes[i] == classes[j]:
                add(f"Duplicate class name: '{classes[i]}'", 1, i)
                add(f"Duplicate class name: '{classes[j]}'", 1, j)

    records = get_record_names(sheet)
    for i in range(0, len(records)-1):
        if records[i] is None: continue
        for j in range(i+1, len(records)):
            if records[i] == records[j]:
                add(f"Duplicate record name: '{records[i]}'", i, 1)
                add(f"Duplicate record name: '{records[j]}'", j, 1)

    for icol in range(2, sheet.max_column + 1):
        if sheet.cell(1, icol).value is not None: continue
        for irow in range(2, sheet.max_row + 1):
            if sheet.cell(irow, icol).value is not None:
                add("Non-empty cell in a column with empty class name.", irow, icol)

    for irow in range(2, sheet.max_row + 1):
        if sheet.cell(irow, 1).value is not None: continue
        for icol in range(2, sheet.max_column + 1):
            if sheet.cell(irow, icol).value is not None:
                add("Non-empty cell in a row with empty record.", irow, icol)

    return result


def merge(destination: Worksheet, source: Worksheet) -> Worksheet:
    """
    Merges source into destination.

    The merge operation appends rows from `source` to `destination`.
    See README.md for details.

    Args:
        destination (Worksheet): A valid (see `validate_input`) LipidQuant Excel sheet into which the `source` is to be merged
        source (Worksheet): A valid (see `validate_input`) LipidQuant Excel sheet to merge into `destination`.

    Returns:
        Worksheet: `destination`

    Notes:
        - `destination` will be modified with the merge (must be writable)
        - the merge operation preserve valid LipidQuant format, so consecutive merges into `destination` are supported
        - the merge operation guarantees that no empty rows or columns remain in the merge (they are ignored in the `source`)
        - the merge operation can be initiated with an empty `destination` sheet
        - `source` remains unmodified (can be read only)
    """
    
    dst_classes = get_class_names(destination)
    src_classes = get_class_names(source)

    def mapper(destination: Worksheet, dst_classes: list[str|None], value: str) -> int|None:
        if str is None: return None
        try:
            # openpyxl Excel cells indexing starts from 1
            # while Python indexing is zero-based.
            # We want this function to obey Excel indexing
            return dst_classes.index(value) + 1 
        except:
            destination.cell(1, destination.max_column+1, value)
            return destination.max_column

    # mapper may already modify destination, see above
    src_dst_map = [mapper(destination, dst_classes, value) for value in src_classes]

    for row in source.iter_rows(2, values_only=True):
       if row[0] is None: continue

       destination.append([row[0]])
       for icol in range(1,len(src_dst_map)):
           if src_dst_map[icol] is None: continue
           destination.cell(destination.max_row, src_dst_map[icol], row[icol])

    return destination

def column_less(sheet: Worksheet, lhs: int, rhs: int) -> bool:
    """
    Compare two columns by their lipid class names.
    (see get_class_names)

    Args:
        sheet (Worksheet): An openpyxl Worksheet object.
        lhs (int): left-hand-side column index (1-based)
        rhs (int): right-hand-side column index (1-based)
    
    Returns:
        bool: True if lhs < rhs

    Warning:
        The `sheet` must contain no empty columns (see `merge`).
    """
    return sheet.cell(1, lhs).value < sheet.cell(1, rhs).value


def column_swap(sheet: Worksheet, i: int, j: int) -> Worksheet:
    """
    Swaps two columns in the sheet.

    In-place swap by swapping every cell value.

    Args:
        sheet (Worksheet): An openpyxl Worksheet object.
        i (int): index of one column to swap (1-based)
        j (int): index of the other column to swap (1-based)

    Returns:
        Worksheet: `sheet` after swap

    Raise:
        InvalidCellIndex
        # swapping in an empty sheet is a no-op; never raise if i, j > 0
    """
    if (

        i > 0 and j > 0 and
        sheet.max_column == 1 and sheet.max_row == 1 and
        sheet.cell(1, 1).value is None

    ): return sheet
        

    if i < 1 or j < 1:
        raise lipiex.InvalidCellIndex("column_swap", min(i, j), 1, sheet.max_column, sheet.max_row)
    if i > sheet.max_column or j > sheet.max_column:
        raise lipiex.InvalidCellIndex("column_swap", max(i, j), 1, sheet.max_column, sheet.max_row)

    if i == j: return sheet

    for row in sheet.iter_rows(values_only=False):
        row[i-1].value, row[j-1].value = row[j-1].value, row[i-1].value

    return sheet

def column_sort(sheet: Worksheet) -> Worksheet:
    """
    Sorts columns in the sheet acc. their lipid class names.
    (see column_less)
    (see column_swap)

    Sorting starts from column 'B' (the first column 'A' is reserved for record names).

    Args:
        sheet (Worksheet): A valid (see validate_input) LipidQuant Excel sheet
    
    Retrurns:
        Worksheet: `sheet` after sort
    """
    return quick_sort(sheet, less=column_less, swap=column_swap, end=sheet.max_column+1, begin=2)

    
def clear_found_values(destination: Worksheet, source: Worksheet) -> Worksheet:
    """
    Check if a value for every record and lipid class in `source` exists in `destination`,
    and if so, clear that value (set it to `None`) in destination.

    The function does not simply search for a value, but it searches for a given
    data record name and lipid class name, and compares the values in `source` and `destination`.

    The idea is that if a sheet M resulted as a merge of sheets A, B and C,
    the consecutive calls to `clear_found_values(M, A)`, `...(M, B)`, `...(M, C)`
    shall result in an empty data set (containing only record names and lipid class names).

    Args:
        destination (Worksheet): A valid (see validate_input) LipidQuant Excel sheet
        source (Worksheet): A valid (see validate_input) LipidQuant Execl sheet to compare to `destination`

    Return:
        Worksheet: `destination`

    Raise:
        Exception: if data entry in `source` is not found in `destination` or the values differ

    Warning:
        `destination` gets modified (do not save it...) 
    """
    dst_records = get_record_names(destination)
    src_records = get_record_names(source)
    dst_classes = get_class_names(destination)
    src_classes = get_class_names(source)

    def mapper(dst_items: list[str|None], value: str) -> int|None:
        if str is None: return None
        try:
            # openpyxl Excel cells indexing starts from 1
            # while Python indexing is zero-based.
            # We want this function to obey Excel indexing
            return dst_items.index(value) + 1
        except:
            raise Exception(f"{value}")

    try:
        src_dst_record_map = [mapper(dst_records, value) for value in src_records]
    except Exception as e:
        raise lipiex.ConsistencyError("Source data not found in destination.", details=[f"Data record: {str(e)}"])
    
    try:
        src_dst_classes_map = [mapper(dst_classes, value) for value in src_classes]
    except Exception as e:
        raise lipiex.ConsistencyError("Source data not found in destination.", details=[f"Lipid class: {str(e)}"])
    
    for src_row in range(2, source.max_row + 1):
        if source.cell(src_row, 1).value is None: continue
        for src_col in range(2, source.max_column + 1):
            src_val = source.cell(src_row, src_col).value
            if src_val is None: continue

            # We index openpyxl cells 1-base, but python list maps 0-based
            dst_row = src_dst_record_map[src_row - 1]
            dst_col = src_dst_classes_map[src_col - 1]
            dst_val = destination.cell(dst_row, dst_col).value

            if dst_val is None:
                raise lipiex.ConsistencyError(
                    "Source data not found in destination.",
                    details=[
                        f"Source record: {source.cell(src_row, 1).value} (row {src_row})",
                        f"Source lipid class: {source.cell(1, src_col).value} (column {src_col})",
                        f"Source value: {src_val}",
                        f"Destination record: {destination.cell(dst_row, 1).value} (row {dst_row})",
                        f"Destination lipid class: {destination.cell(1, dst_col).value} (column {dst_col})",
                        f"Destination value: [empty cell]"
                    ]
                )

            if dst_val != src_val:
                raise lipiex.ConsistencyError(
                    "Source data differs for destination.",
                    details=[
                        f"Source record: {source.cell(src_row, 1).value} (row {src_row})",
                        f"Source lipid class: {source.cell(1, src_col).value} (column {src_col})",
                        f"Source value: {src_val}",
                        f"Destination record: {destination.cell(dst_row, 1).value} (row {dst_row})",
                        f"Destination lipid class: {destination.cell(1, dst_col).value} (column {dst_col})",
                        f"Destination value: {dst_val}",
                    ]
                )

            destination.cell(dst_row, dst_col).value = None

    return destination


def has_empty_data_set(sheet: Worksheet) -> bool:
    """
    Check whether all values in the dataset are empty cells.
    (see clear_found_values)

    Args:
        sheet (Worksheet): A valid (see validate_input) LilpidQuant Excel sheet
    
    Return:
        bool: True if the sheet is empty up to the record names and lipid class names
              (see get_record_names)
              (see get_class_names)
    """
    for row in sheet.iter_rows(min_row=2, values_only=True):
        for col in row[1:]: # row index is now 0-based
            if col is not None: return False

    return True

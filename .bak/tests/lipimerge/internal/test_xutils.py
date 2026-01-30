import pytest
from contextlib import nullcontext as does_not_raise
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
import lipimerge.internal.xutils as xutils
import lipimerge.internal.exceptions as lipiex

# ============================================================================
# Openpyxl Assumptions

def test_openpyxl_empty_sheet_has_max_col_and_row_1():
    with xutils.workbook('./tests/data/openpyxl.xlsx') as wb:
        sheet = wb["Empty"]
        assert sheet.max_column == 1 and sheet.max_row == 1


def test_openpyxl_empty_sheet_yields_one_None_value():
    with xutils.workbook('./tests/data/openpyxl.xlsx') as wb:
        sheet = wb["Empty"]
        with does_not_raise():
            row_set = [value for value in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
            col_set = [value for value in next(sheet.iter_cols(min_col=1, max_col=1, values_only=True))]

        assert len(row_set) == 1 and len(col_set) == 1 and row_set[0] is None and col_set[0] is None


def test_openpyxl_creates_cell_if_not_exist():
    with xutils.workbook('./tests/data/openpyxl.xlsx') as wb:
        sheet = wb["Empty"]
        sheet.cell(3, 3, "x")
        assert sheet.max_column == 3 and sheet.max_row == 3 and xutils.equal(sheet, wb["AddCell"])


def test_openpyxl_iter_rows_start_from_beyond_max_rows_does_not_raise():
    with xutils.workbook('./tests/data/openpyxl.xlsx') as wb:
        sheet = wb["IterRows"]
        with does_not_raise():
            result = [value for value in sheet.iter_rows(min_row=sheet.max_row + 1, max_row=sheet.max_row, values_only=True)]
            assert result == []


# ============================================================================
# Get Classes

def test_get_class_names_empty_returns_single_None():
    with xutils.workbook('./tests/data/get_classes.xlsx') as wb:
        classes = xutils.get_class_names(wb["Empty"])
        assert classes == [None]


def test_get_class_names_header_only():
    with xutils.workbook('./tests/data/get_classes.xlsx') as wb:
        classes = xutils.get_class_names(wb["HeaderOnly"])
        assert classes == [None, "B", "C", None, "E"]


def test_get_class_names_AColumn_only():
    with xutils.workbook('./tests/data/get_classes.xlsx') as wb:
        classes = xutils.get_class_names(wb["AColumnOnly"])
        assert classes == [None]

# ============================================================================
# Get Records

def test_get_record_names_empty_returns_single_None():
    with xutils.workbook('./tests/data/get_records.xlsx') as wb:
        records = xutils.get_record_names(wb["Empty"])
        assert records == [None]


def test_get_record_names_records():
    with xutils.workbook('./tests/data/get_records.xlsx') as wb:
        records = xutils.get_record_names(wb["HeaderOnly"])
        assert records == [None]


def test_get_record_names_AColumn_only():
    with xutils.workbook('./tests/data/get_records.xlsx') as wb:
        records = xutils.get_record_names(wb["AColumnOnly"])
        assert records == [None, "I", "J", None, "L"]

# ============================================================================
# Validate Input

@pytest.mark.parametrize("test_sheet", [
    ("Empty"),
    ("SingleRow"),
    ("SingleColumn"),
    ("ValidNoEmpty"),
    ("ValidEmptyColumns"),
    ("ValidEmptyRows"),
])
def test_validate_input_ok(test_sheet: Worksheet):

    with xutils.workbook('./tests/data/input_validation.xlsx') as wb:
        sheet = wb[test_sheet]
        result = xutils.validate_input(sheet)

        msg = ""
        for s in result: msg += f"{str(s)}; "
        assert len(result) == 0, msg


@pytest.mark.parametrize("test_sheet", [
    ("A1NotEmpty"),
    ("InvalidColumn-1"),
    ("InvalidColumn-2"),
    ("InvalidColumn-3"),
    ("InvalidRow-1"),
    ("InvalidRow-2"),
    ("InvalidRow-3"),
    ("DuplicateRow-1"),
    ("DuplicateRow-2"),
    ("DuplicateRow-3"),
    ("DuplicateRow-4"),
    ("DuplicateRow-5"),
    ("DuplicateRow-6"),
    ("DuplicateCol-1"),
    ("DuplicateCol-2"),
    ("DuplicateCol-3"),
    ("DuplicateCol-4"),
    ("DuplicateCol-5"),
    ("DuplicateCol-6"),
])
def test_validate_input_nok(test_sheet: Worksheet):

    with xutils.workbook('./tests/data/input_validation.xlsx') as wb:
        sheet = wb[test_sheet]
        result = xutils.validate_input(sheet)

        msg = ""
        for s in result: msg += f"{str(s)}; "
        assert len(result) != 0, msg


# ============================================================================
# Merge

def test_merge():

    destination = Workbook()
    result = destination.active
    with xutils.workbook('./tests/data/merge.xlsx') as wb:
        for i in range(1, len(wb.sheetnames)//2 + 1):
            xutils.merge(result, wb[f"Data{i}"])
            if not xutils.equal(result, wb[f"Merge{i}"]): 
                destination.save('./tests/data/debug.xlsx')
                pytest.fail(f"Merge {i} failure.")
        
    assert True


# ============================================================================
# Column Less

@pytest.mark.parametrize("i, j, expected", [
    (2, 3, True),
    (3, 2, False),
    (2, 4, True),
    (4, 2, False),
    (3, 4, False),
    (4, 3, True),
    (2, 2, False),
])
def test_column_less(i: int, j: int, expected: bool):

    with xutils.workbook('./tests/data/column_less.xlsx') as wb:
        sheet = wb["Less"]
        assert xutils.column_less(sheet, i, j) == expected

# ============================================================================
# Column Swap

@pytest.mark.parametrize("i, j", [
    (1, 2),
    (2, 1),
    (2, 4),
    (4, 2),
    (1, 1),
    (5, 5),
])
def test_column_swap_empty_sheet_does_not_raise(i: int, j: int):

    with xutils.workbook('./tests/data/column_swap.xlsx') as wb:
        sheet = wb["Empty"]
        with does_not_raise(): xutils.column_swap(sheet, i, j)


@pytest.mark.parametrize("i, j", [
    (0, 0),
    (-1, 1),
    (0, 1),
    (1, 0),
    (-1, 1),
    (1, -1),
])
def test_column_swap_empty_sheet_raises_InvalidCellIndex_if_index_below_1(i: int, j: int):

    with xutils.workbook('./tests/data/column_swap.xlsx') as wb:

        sheet = wb["Empty"]

        try:
            xutils.column_swap(sheet, i, j)
        except Exception as e:
            assert isinstance(e, lipiex.InvalidCellIndex), f"Expected `InvalidCellIndex` but got `{type(e).__name__}` with message '{e}'."
        else:
            pytest.fail("Expected `InvalidCellIndex` but no exception was raised.")


@pytest.mark.parametrize("i, j", [
    (2, 5),
    (5, 2),
    (0, 1),
    (1, 0),
    (-1, 1),
    (1, -1),
])
def test_column_swap_nonempty_sheet_raises_InvalidCellIndex(i: int, j: int):

    with xutils.workbook('./tests/data/column_swap.xlsx') as wb:

        sheet = wb["Nonempty"]

        try:
            xutils.column_swap(sheet, i, j)
        except Exception as e:
            assert isinstance(e, lipiex.InvalidCellIndex), f"Expected `InvalidCellIndex` but got `{type(e).__name__}` with message '{e}'."
        else:
            pytest.fail("Expected `InvalidCellIndex` but no exception was raised.")


@pytest.mark.parametrize("i, j", [
    (1, 2),
    (2, 1),
    (2, 4),
    (4, 2),
])
def test_column_swap(i: int, j: int):

    with xutils.workbook('./tests/data/column_swap.xlsx') as wb:
        sheet = wb["Nonempty"]
        xutils.column_swap(sheet, i, j)
        assert xutils.equal(sheet, wb[f"Swap{min(i,j)}-{max(i,j)}"])


@pytest.mark.parametrize("i, j", [
    (1, 3),
    (3, 1),
])
def test_column_swap_empty_column(i: int, j: int):

    with xutils.workbook('./tests/data/column_swap.xlsx') as wb:
        sheet = wb["EmptyColumn"]
        xutils.column_swap(sheet, i, j)
        assert xutils.equal(sheet, wb["SwapEmptyColumn"])


@pytest.mark.parametrize("i, j", [
    (1, 3),
    (3, 1),
    (2, 4),
    (4, 2),
])
def test_column_swap_varying_heights(i: int, j: int):

    with xutils.workbook('./tests/data/column_swap.xlsx') as wb:
        sheet = wb["VaryingHeights"]
        xutils.column_swap(sheet, i, j)
        assert xutils.equal(sheet, wb[f"SwapVaryingHeights{min(i,j)}-{max(i,j)}"])


# ============================================================================
# Column Sort

def test_column_sort_empty_does_not_raise():

    with xutils.workbook('./tests/data/column_sort.xlsx') as wb:
        sheet = wb["Empty"]
        with does_not_raise():
            assert xutils.equal(xutils.column_sort(sheet), wb["Empty Sorted"])


def test_column_sort_single_column_does_not_raise():

    with xutils.workbook('./tests/data/column_sort.xlsx') as wb:
        sheet = wb["SingleColumn"]
        with does_not_raise():
            assert xutils.equal(xutils.column_sort(sheet), wb["SingleColumn Sorted"])
                                                                    

def test_column_sort():

    with xutils.workbook('./tests/data/column_sort.xlsx') as wb:
        sheet = wb["Sort"]
        assert xutils.equal(xutils.column_sort(sheet), wb["Sorted"])


# ============================================================================
# Clear Found Values

def reproduce_sheet(sheet: Worksheet) -> Worksheet:
    result = Workbook().active
    for i in range(1, sheet.max_row + 1):
        for j in range(1, sheet.max_column+ 1):
            result.cell(i, j, value=sheet.cell(i, j).value)

    return result


@pytest.mark.parametrize("sheet", [
    ("FullData"),
    ("PartialData"),
    ("BorderData"),
])
def test_clear_found_values(sheet):
    with xutils.workbook('./tests/data/clear_found_values.xlsx') as wb:
        copy = reproduce_sheet(wb["FullData"])
        xutils.clear_found_values(copy, wb[sheet])
        assert xutils.equal(copy, wb[f"{sheet} Clear"])    


@pytest.mark.parametrize("sheet", [
    ("ErrorExtraColumn"),
    ("ErrorExtraRow"),
    ("ErrorInvalidValue-1"),
    ("ErrorInvalidValue-2"),
    ("ErrorInvalidValue-3"),
    ("ErrorMissingValue-1"),
    ("ErrorMissingValue-2"),
    ("ErrorMissingValue-3"),
])
def test_clear_found_values_raises_on_error(sheet):
    with xutils.workbook('./tests/data/clear_found_values.xlsx') as wb:
        copy = reproduce_sheet(wb["FullData"])
        with pytest.raises(Exception):
            xutils.clear_found_values(copy, wb[sheet])
            

# ============================================================================
# Has Empty Data Sheet

def test_has_empty_data_set():
    with xutils.workbook('./tests/data/has_empty_data_set.xlsx') as wb:
        assert xutils.has_empty_data_set(wb["Empty Data Set"]) == True


@pytest.mark.parametrize("sheet", [
    ("Nonempty 1"),
    ("Nonempty 2"),
    ("Nonempty 3"),
])
def test_has_empty_data_set_reveals_nonempty(sheet):
    with xutils.workbook('./tests/data/has_empty_data_set.xlsx') as wb:
        assert xutils.has_empty_data_set(wb[sheet]) == False
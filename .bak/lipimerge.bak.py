# IMPORTS

# .. Global

import sys
from openpyxl import load_workbook

# .. Local

# MAIN


# Bubble sort >

# Bubble sort columns in an Excel sheet based on the first row values
# This script sorts columns in an Excel sheet by comparing the values in the first row of each column.
# It uses a bubble sort algorithm to swap columns until they are sorted in ascending order. 
from openpyxl import load_workbook

def main():
    wb = load_workbook('Data/TestFile-1.xlsx', data_only=True)
    sheet = wb['Conc']  # or wb.active

    # Iterate over rows and print plain cell values
    for row in sheet.iter_cols(values_only=True):
        print(row[0], end=' ')
        for cell in row:
            if cell is not None:
                print(cell, end=' ')
            else:
                print('None', end=' ')
        print()

if __name__ == "__main__":
    main()
